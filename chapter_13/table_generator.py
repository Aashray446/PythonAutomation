from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
wb = Workbook()
ws = wb.active
ws.title = "Multiplication table"

def add(n):
    for i in range(1,n+1):
        for j in range(1,n+1):
            char = get_column_letter(j)
            ws[char+str(i)] = i*j

def style_row(n):
    count = 1
    for i in range(2,n+2):
        ws['A'+str(i)] = count
        ws['A'+str(i)].font = Font(bold=True)
        char = chr(64+i)
        ws[char+"1"] = count
        ws[char+"1"].font = Font(bold=True) 
        count = count + 1

num = int(input("Enter the Number: "))

try:
    add(num)
    ws.insert_cols(1)
    ws.insert_rows(1)
    style_row(num)
    wb.save('./multi_table.xlsx')
    print("Operation Succesfull")
except:
    print("Oops! There are some errors")