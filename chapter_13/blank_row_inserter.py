from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def insert(n,m, file_name):
    wb = load_workbook(file_name)
    ws = wb.active
    for row in range(1,m):
        ws.insert_rows(n)
    wb.save(file_name)

 #Taking Inputs   
n = int(input("At what row: "))
m = int(input("How many times: "))
file_name = input("Enter file name")

#Trying to perform the operation
try:
    insert(n,m)
    print("Operation is successfull")
except:
    print("Either file not found or there are some unexpected errors")
