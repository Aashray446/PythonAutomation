import openpyxl


wb = openpyxl.load_workbook('multi_table.xlsx')
sheet = wb.active

text = []
full_list = []

for col in range(1, sheet.max_column + 1):
    for row in range(1, sheet.max_row + 1):
        text.append(sheet.cell(row=row, column=col).value)
        text = [x for x in text if x is not None]
    full_list.append(text)
    text = []


col = 1
for list in full_list:
    text_file = open('spread_to_text' + str(col) + '.txt', 'w')
    for item in list:
        text_file.write(str(item) + '\n')
    col += 1
    text_file.close()