
import openpyxl
import os
from openpyxl.utils import get_column_letter


wb = openpyxl.Workbook()
sheet = wb.active


text_list = []
for text in os.listdir('.'):
    if text.endswith('.txt'):
        text = open(text)
        text_list.append(text.readlines())



col = 1
for text in text_list:
    row = 1
    list_length = []
    for sentence in text:
   
        sheet.cell(row=row, column=col).value = sentence
        row += 1
 
        length = len(sentence)
        list_length.append(length)
        max_length = max(list_length)
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = max_length

    col += 1

wb.save('text_to_spreadsheet.xlsx')
